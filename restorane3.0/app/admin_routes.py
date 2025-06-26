from fastapi import APIRouter, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from app.database import get_connection
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import csv
import io

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


@router.get("/admin/orders", response_class=HTMLResponse)
def admin_orders(request: Request):
    if request.session.get("role") != "admin":
        return RedirectResponse("/login", status_code=303)

    # Параметры фильтрации
    start_date = request.query_params.get("start")
    end_date = request.query_params.get("end")
    waiter_id = request.query_params.get("waiter_id")
    status = request.query_params.get("status")

    with get_connection() as conn:
        cur = conn.cursor()

        # Основной запрос для заказов
        orders_query = """
            SELECT 
                o.id, o.created_at, o.status, o.total_price,
                u.username as client_name,
                w.username as waiter_name,
                t.table_number
            FROM orders o
            JOIN users u ON o.user_id = u.id
            JOIN users w ON o.waiter_id = w.id
            JOIN restaurant_tables t ON o.table_id = t.id
        """
        
        conditions = []
        params = []
        
        if start_date:
            conditions.append("o.created_at::date >= %s")
            params.append(start_date)
        if end_date:
            conditions.append("o.created_at::date <= %s")
            params.append(end_date)
        if waiter_id:
            conditions.append("o.waiter_id = %s")
            params.append(waiter_id)
        if status:
            conditions.append("o.status = %s")
            params.append(status)
            
        if conditions:
            orders_query += " WHERE " + " AND ".join(conditions)
            
        orders_query += " ORDER BY o.created_at DESC"
        
        cur.execute(orders_query, params)
        orders = [
            {
                "id": row[0],
                "created_at": row[1],
                "status": row[2],
                "total_price": float(row[3] or 0),
                "client_name": row[4],
                "waiter_name": row[5],
                "table_number": row[6]
            }
            for row in cur.fetchall()
        ]

        # Получаем список официантов для фильтра
        cur.execute("SELECT id, username FROM users WHERE role = 'waiter'")
        waiters = [{"id": row[0], "name": row[1]} for row in cur.fetchall()]

        # Статусы заказов для фильтра
        statuses = ["невыполнен", "в процессе", "завершен", "отменен"]

    return templates.TemplateResponse("admin_orders.html", {
        "request": request,
        "orders": orders,
        "waiters": waiters,
        "statuses": statuses,
        "filters": {
            "start": start_date,
            "end": end_date,
            "waiter_id": waiter_id,
            "status": status
        }
    })

@router.get("/admin/order/{order_id}", response_class=HTMLResponse)
def order_details(request: Request, order_id: int):
    if request.session.get("role") != "admin":
        return RedirectResponse("/login", status_code=303)

    with get_connection() as conn:
        cur = conn.cursor()

        # Информация о заказе
        cur.execute("""
            SELECT 
                o.id, o.created_at, o.status, o.total_price,
                u.username as client_name,
                w.username as waiter_name,
                t.table_number
            FROM orders o
            JOIN users u ON o.user_id = u.id
            JOIN users w ON o.waiter_id = w.id
            JOIN restaurant_tables t ON o.table_id = t.id
            WHERE o.id = %s
        """, (order_id,))
        order = cur.fetchone()
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
            
        order_data = {
            "id": order[0],
            "created_at": order[1],
            "status": order[2],
            "total_price": float(order[3] or 0),
            "client_name": order[4],
            "waiter_name": order[5],
            "table_number": order[6]
        }

        # Позиции заказа
        cur.execute("""
            SELECT 
                mi.name, oi.quantity, mi.price,
                (oi.quantity * mi.price) as total
            FROM order_items oi
            JOIN menu_items mi ON oi.menu_item_id = mi.id
            WHERE oi.order_id = %s
        """, (order_id,))
        items = [
            {
                "name": row[0],
                "quantity": row[1],
                "price": float(row[2]),
                "total": float(row[3])
            }
            for row in cur.fetchall()
        ]

    return templates.TemplateResponse("admin_order_details.html", {
        "request": request,
        "order": order_data,
        "items": items
    })

@router.get("/admin/reservation_stats", response_class=HTMLResponse)
def reservation_stats(request: Request):
    if request.session.get("role") != "admin":
        return RedirectResponse("/login", status_code=303)

    with get_connection() as conn:
        cur = conn.cursor()

        # Получаем список всех столов
        cur.execute("SELECT id, table_number FROM restaurant_tables ORDER BY table_number")
        tables = [{"id": row[0], "number": row[1]} for row in cur.fetchall()]

        # Получаем статистику бронирований по дням
        cur.execute("""
            SELECT 
                DATE(reservation_time) as date,
                table_id,
                COUNT(*) as count,
                t.table_number
            FROM reservations r
            JOIN restaurant_tables t ON r.table_id = t.id
            GROUP BY DATE(reservation_time), table_id, t.table_number
            ORDER BY date DESC, t.table_number
        """)
        stats = cur.fetchall()

        # Формируем данные для таблицы
        dates = sorted({row[0].strftime("%Y-%m-%d") for row in stats}, reverse=True)
        stats_dict = {(row[0].strftime("%Y-%m-%d"), row[3]): row[2] for row in stats}

    return templates.TemplateResponse("reservation_stats.html", {
        "request": request,
        "dates": dates,
        "tables": [t["number"] for t in tables],
        "stats_dict": stats_dict
    })

# Add these routes to your admin section
@router.get("/admin/inventory", response_class=HTMLResponse)
async def inventory_management(request: Request):
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Get all menu items with their inventory data
        cur.execute("""
            SELECT m.id, m.name, m.price, 
                   COALESCE(di.current_quantity, 0) as current_quantity,
                   COALESCE(di.max_quantity, 0) as max_quantity
            FROM menu_items m
            LEFT JOIN dish_inventory di ON m.id = di.menu_item_id
            ORDER BY m.id
        """)
        items = cur.fetchall()
    
    return templates.TemplateResponse("admin_inventory.html", {
        "request": request,
        "items": items
    })
@router.post("/admin/inventory/restock_all")
async def restock_all(request: Request):
    with get_connection() as conn:
        cur = conn.cursor()
        
        try:
            cur.execute("""
                UPDATE dish_inventory
                SET current_quantity = max_quantity
            """)
            conn.commit()
            request.session["flash_message"] = "All items restocked to max quantity!"
            request.session["flash_type"] = "success"
        except Exception as e:
            conn.rollback()
            request.session["flash_message"] = f"Error during restock: {str(e)}"
            request.session["flash_type"] = "danger"
    
    return RedirectResponse(url="/admin/inventory", status_code=303)

@router.post("/admin/inventory/reset_all")
async def reset_all(request: Request):
    with get_connection() as conn:
        cur = conn.cursor()
        
        try:
            cur.execute("""
                UPDATE dish_inventory
                SET current_quantity = 0
            """)
            conn.commit()
            request.session["flash_message"] = "All quantities reset to zero!"
            request.session["flash_type"] = "warning"
        except Exception as e:
            conn.rollback()
            request.session["flash_message"] = f"Error during reset: {str(e)}"
            request.session["flash_type"] = "danger"
    
    return RedirectResponse(url="/admin/inventory", status_code=303)
@router.post("/admin/inventory/update")
async def update_inventory(
    request: Request,
    menu_item_id: int = Form(...),
    current_quantity: int = Form(...),
    max_quantity: int = Form(...)
):
    with get_connection() as conn:
        cur = conn.cursor()
        
        try:
            # Update or insert inventory record
            cur.execute("""
                INSERT INTO dish_inventory (menu_item_id, current_quantity, max_quantity)
                VALUES (%s, %s, %s)
                ON CONFLICT (menu_item_id) DO UPDATE
                SET current_quantity = EXCLUDED.current_quantity,
                    max_quantity = EXCLUDED.max_quantity
            """, (menu_item_id, current_quantity, max_quantity))
            
            conn.commit()
            request.session["flash_message"] = "Inventory updated successfully!"
            request.session["flash_type"] = "success"
        except Exception as e:
            conn.rollback()
            request.session["flash_message"] = f"Error updating inventory: {str(e)}"
            request.session["flash_type"] = "danger"
    
    return RedirectResponse(url="/admin/inventory", status_code=303)
@router.get("/admin/reservation_stats/excel")
def export_reservation_stats():
    with get_connection() as conn:
        cur = conn.cursor()

        # Получаем данные для Excel
        cur.execute("""
            SELECT 
                DATE(r.reservation_time) as date,
                t.table_number,
                COUNT(*) as count
            FROM reservations r
            JOIN restaurant_tables t ON r.table_id = t.id
            GROUP BY DATE(r.reservation_time), t.table_number
            ORDER BY date DESC, t.table_number
        """)
        stats = cur.fetchall()

        # Получаем список всех столов
        cur.execute("SELECT table_number FROM restaurant_tables ORDER BY table_number")
        tables = [row[0] for row in cur.fetchall()]

    # Создаем Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика бронирований"

    # Стили для Excel
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                  top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Заголовки
    ws.append(["Дата"] + [f"Стол №{num}" for num in tables])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Данные
    current_date = None
    for date, table_num, count in stats:
        date_str = date.strftime("%Y-%m-%d")
        
        if date_str != current_date:
            ws.append([date_str] + [0] * len(tables))
            current_date = date_str
        
        col_idx = tables.index(table_num) + 2
        ws.cell(row=ws.max_row, column=col_idx, value=count)

    # Форматирование
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = center_alignment
            if cell.column == 1:  # Дата
                cell.font = Font(bold=True)
            elif cell.value:  # Число бронирований
                if cell.value >= 3:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                elif cell.value == 2:
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                elif cell.value == 1:
                    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    # Настройка ширины столбцов
    for i in range(2, len(tables) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.column_dimensions['A'].width = 12

    # Сохраняем в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reservation_stats.xlsx"}
    )

@router.get("/admin", response_class=HTMLResponse)
def admin_dashboard(request: Request):
    if request.session.get("role") != "admin":
        return RedirectResponse("/login", status_code=303)

    start_date = request.query_params.get("start")
    end_date = request.query_params.get("end")

    with get_connection() as conn:
        cur = conn.cursor()

        # Получаем список пользователей
        cur.execute("SELECT id, username, role FROM users")
        users = [dict(id=r[0], username=r[1], role=r[2]) for r in cur.fetchall()]

        # Получаем бронирования
        cur.execute("""
            SELECT r.id, u.username, r.table_id, r.reservation_time, r.status
            FROM reservations r
            JOIN users u ON r.user_id = u.id
            ORDER BY r.reservation_time DESC
        """)
        reservations = [
            dict(id=r[0], username=r[1], table_id=r[2], time=r[3], status=r[4])
            for r in cur.fetchall()
        ]

        # Получаем список столов
        cur.execute("SELECT id, table_number, seats FROM restaurant_tables")
        tables = [dict(id=r[0], number=r[1], seats=r[2]) for r in cur.fetchall()]

        # Статистика заказов по дням
        query = """
            SELECT DATE(created_at), COUNT(*), SUM(total_price)
            FROM orders
        """
        params = []
        if start_date and end_date:
            query += " WHERE created_at::date BETWEEN %s AND %s"
            params = [start_date, end_date]
        query += " GROUP BY DATE(created_at) ORDER BY DATE(created_at) DESC"

        cur.execute(query, params)
        order_stats = [
            {"date": row[0].strftime("%Y-%m-%d"), "total_orders": row[1], "total_revenue": float(row[2] or 0)}
            for row in cur.fetchall()
        ]

        # Статистика по официантам
        waiter_query = """
            SELECT 
                u.id as waiter_id,
                u.username as waiter_name,
                COUNT(o.id) as total_orders,
                SUM(o.total_price) as total_revenue,
                AVG(o.total_price) as avg_order_value
            FROM orders o
            JOIN users u ON o.waiter_id = u.id
        """
        waiter_params = []
        if start_date and end_date:
            waiter_query += " WHERE o.created_at::date BETWEEN %s AND %s"
            waiter_params = [start_date, end_date]
        waiter_query += " GROUP BY u.id, u.username ORDER BY total_revenue DESC"

        cur.execute(waiter_query, waiter_params)
        waiter_stats = [
            {
                "waiter_id": row[0],
                "waiter_name": row[1],
                "total_orders": row[2],
                "total_revenue": float(row[3] or 0),
                "avg_order_value": float(row[4] or 0)
            }
            for row in cur.fetchall()
        ]

    return templates.TemplateResponse("admin_dashboard.html", {
        "request": request,
        "users": users,
        "reservations": reservations,
        "tables": tables,
        "order_stats": order_stats,
        "waiter_stats": waiter_stats
    })


@router.post("/admin/delete_reservation")
def delete_reservation(request: Request, reservation_id: int = Form(...)):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM reservations WHERE id = %s", (reservation_id,))
        conn.commit()
    return RedirectResponse("/admin", status_code=303)


@router.post("/admin/create_reservation")
def add_reservation(
    request: Request,
    user_id: int = Form(...),
    table_id: int = Form(...),
    reservation_time: str = Form(...)
):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO reservations (user_id, table_id, reservation_time, status, created_at)
            VALUES (%s, %s, %s, %s, %s)
        """, (user_id, table_id, reservation_time, "active", datetime.now()))
        conn.commit()
    return RedirectResponse("/admin", status_code=303)


@router.get("/admin/export_csv")
def export_csv():
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT DATE(created_at), COUNT(*), SUM(total_price)
            FROM orders
            GROUP BY DATE(created_at)
            ORDER BY DATE(created_at) DESC
        """)
        data = cur.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Дата", "Количество заказов", "Сумма"])
    for row in data:
        writer.writerow([row[0].strftime("%Y-%m-%d"), row[1], row[2]])

    output.seek(0)
    return StreamingResponse(output, media_type="text/csv", headers={
        "Content-Disposition": "attachment; filename=order_statistics.csv"
    })
